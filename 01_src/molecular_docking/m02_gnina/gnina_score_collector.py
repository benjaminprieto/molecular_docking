#!/usr/bin/env python3
"""
GNINA Score Collector - Core Module (02c)
===========================================
Collects, enriches, and exports GNINA docking scores.

Reads gnina_results.csv from 02b, merges with molecule metadata from 00a,
ranks by multiple criteria (Vina affinity, CNN score, CNN affinity),
and produces final CSV + Excel outputs.

Does NOT perform cross-engine comparison (GNINA vs DOCK6) —
that responsibility belongs to dock2profile.

Location: 01_src/molecular_docking/m02_gnina/gnina_score_collector.py

Project: molecular_docking
Module: 02c (core)
Version: 1.1 — keep_all_poses for clustering (2026-03-21)
"""

import logging
import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Any, Union

import pandas as pd
import numpy as np

logger = logging.getLogger(__name__)

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# =============================================================================
# ALL-POSES EXTRACTION FROM GNINA OUTPUT SDF FILES
# =============================================================================

def parse_gnina_sdf_all_poses(sdf_path: str, name: str) -> List[Dict[str, Any]]:
    """
    Parse a GNINA output SDF file and extract scores for ALL poses.

    GNINA writes per-pose properties:
      - minimizedAffinity  → vina_affinity
      - CNNscore           → cnn_score
      - CNNaffinity        → cnn_affinity
      - CNN_VS             → cnn_vs (optional)

    Args:
        sdf_path: Path to {name}_gnina.sdf
        name: Molecule name

    Returns:
        List of dicts, one per pose
    """
    path = Path(sdf_path)
    if not path.exists() or path.stat().st_size == 0:
        return []

    content = path.read_text()
    blocks = content.split("$$$$")
    poses = []

    for idx, block in enumerate(blocks):
        block = block.strip()
        if not block:
            continue

        pose = {"Name": name, "pose_index": idx}

        # Extract properties using regex
        m = re.search(r"<minimizedAffinity>\s*\n\s*([-\d.]+)", block)
        if m:
            pose["vina_affinity"] = float(m.group(1))

        m = re.search(r"<CNNscore>\s*\n\s*([\d.]+)", block)
        if m:
            pose["cnn_score"] = float(m.group(1))

        m = re.search(r"<CNNaffinity>\s*\n\s*([\d.]+)", block)
        if m:
            pose["cnn_affinity"] = float(m.group(1))

        m = re.search(r"<CNN_VS>\s*\n\s*([\d.]+)", block)
        if m:
            pose["cnn_vs"] = float(m.group(1))

        m = re.search(r"<CNNaffinity_variance>\s*\n\s*([\d.]+)", block)
        if m:
            pose["cnn_affinity_variance"] = float(m.group(1))

        # Only include if we got at least one score
        if any(k in pose for k in ("vina_affinity", "cnn_score", "cnn_affinity")):
            poses.append(pose)

    return poses


# =============================================================================
# RANKING
# =============================================================================

def enrich_and_rank(
        gnina_df: pd.DataFrame,
        meta_df: Optional[pd.DataFrame] = None,
        name_column: str = "Name",
        top_n: int = 10,
        rank_by: str = "cnn_affinity",
) -> pd.DataFrame:
    """
    Merge GNINA scores with molecule metadata and compute rankings.

    Ranks by three criteria:
      - Vina affinity (lower = better)
      - CNN score (higher = better)
      - CNN affinity (lower = better)

    Args:
        gnina_df:    DataFrame from 02b gnina_results.csv
        meta_df:     DataFrame from 00a unique_molecules.csv (optional)
        name_column: Name column in metadata CSV
        top_n:       Number of top molecules to flag
        rank_by:     Primary ranking column ("cnn_affinity", "cnn_score", "vina_affinity")
    """
    df = gnina_df.copy()

    # Merge with metadata
    if meta_df is not None and name_column in meta_df.columns:
        df['name'] = df['name'].astype(str)
        meta_df[name_column] = meta_df[name_column].astype(str)
        merge_cols = [c for c in meta_df.columns if c not in df.columns or c == name_column]
        df = df.merge(meta_df[merge_cols], left_on='name', right_on=name_column, how='left')
        if name_column != 'name' and name_column in df.columns:
            df.drop(columns=[name_column], inplace=True, errors='ignore')

    successful = df['success'] == True

    # Rank by Vina affinity (lower = better → ascending)
    if 'vina_affinity' in df.columns:
        df['Vina_Rank'] = np.nan
        df.loc[successful, 'Vina_Rank'] = (
            df.loc[successful, 'vina_affinity'].rank(method='min', ascending=True).astype(int)
        )

    # Rank by CNN score (higher = better → descending)
    if 'cnn_score' in df.columns:
        df['CNN_Score_Rank'] = np.nan
        df.loc[successful, 'CNN_Score_Rank'] = (
            df.loc[successful, 'cnn_score'].rank(method='min', ascending=False).astype(int)
        )

    # Rank by CNN affinity (lower = better → ascending)
    if 'cnn_affinity' in df.columns:
        df['CNN_Affinity_Rank'] = np.nan
        df.loc[successful, 'CNN_Affinity_Rank'] = (
            df.loc[successful, 'cnn_affinity'].rank(method='min', ascending=True).astype(int)
        )

    # Consensus rank (average of three ranks)
    rank_cols = [c for c in ['Vina_Rank', 'CNN_Score_Rank', 'CNN_Affinity_Rank'] if c in df.columns]
    if rank_cols:
        df['Consensus_Rank'] = df[rank_cols].mean(axis=1)
        df.loc[successful, 'Consensus_Rank'] = (
            df.loc[successful, 'Consensus_Rank'].rank(method='min', ascending=True).astype(int)
        )

    # Primary rank for sorting
    primary_rank = {
        "cnn_affinity": "CNN_Affinity_Rank",
        "cnn_score": "CNN_Score_Rank",
        "vina_affinity": "Vina_Rank",
        "consensus": "Consensus_Rank",
    }.get(rank_by, "Consensus_Rank")

    # Flag top N
    df['Is_Top_N'] = False
    if primary_rank in df.columns:
        df.loc[df[primary_rank] <= top_n, 'Is_Top_N'] = True

    # Flag reference
    df['Is_Reference'] = False
    if 'is_reference' in df.columns:
        df['Is_Reference'] = df['is_reference'].fillna(False).astype(bool)
    elif 'Strategy' in df.columns:
        df.loc[df['Strategy'] == 'reference_control', 'Is_Reference'] = True

    # Sort
    if primary_rank in df.columns:
        df = df.sort_values(primary_rank, na_position='last').reset_index(drop=True)

    return df


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def export_to_excel(df: pd.DataFrame, output_path: str, sheet_name: str = "GNINA_Scores") -> bool:
    """Export scores to formatted Excel."""
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available — skipping Excel export")
        return False

    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        headers = list(df.columns)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, (col_name, value) in enumerate(row.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if pd.isna(value):
                    cell.value = None
                elif isinstance(value, (np.integer,)):
                    cell.value = int(value)
                elif isinstance(value, (np.floating,)):
                    cell.value = float(value)
                elif isinstance(value, (np.bool_,)):
                    cell.value = bool(value)
                else:
                    cell.value = value

                if col_name in ('vina_affinity', 'cnn_affinity') and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
                elif col_name == 'cnn_score' and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.000'

            if row.get('Is_Top_N', False):
                green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = green

            if row.get('Is_Reference', False):
                yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow

        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row_idx in range(2, min(len(df) + 2, 50)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

        ws.freeze_panes = "A2"
        wb.save(output_path)
        logger.info(f"  Saved Excel: {output_path}")
        return True
    except Exception as e:
        logger.error(f"Excel export failed: {e}")
        return False


# =============================================================================
# SUMMARY
# =============================================================================

def generate_summary(df: pd.DataFrame, output_path: str, top_n: int = 10, rank_by: str = "cnn_affinity") -> str:
    w = 78
    successful = df[df['success'] == True]
    n_ok = len(successful)

    lines = [
        "=" * w,
        "02c GNINA SCORE COLLECTION — SUMMARY",
        "=" * w,
        "",
        f"Total molecules:   {len(df)}",
        f"Successful:        {n_ok}",
        f"Failed:            {len(df) - n_ok}",
        f"Ranked by:         {rank_by}",
    ]

    if n_ok > 0:
        for col, label in [('vina_affinity', 'Vina Affinity'), ('cnn_score', 'CNN Score'), ('cnn_affinity', 'CNN Affinity')]:
            if col in successful.columns:
                vals = successful[col].dropna()
                if not vals.empty:
                    lines.extend([
                        "",
                        f"{label.upper()} STATISTICS:",
                        f"  Best:    {vals.min():.3f}" if 'score' in col else f"  Best:    {vals.min():.2f}",
                        f"  Worst:   {vals.max():.3f}" if 'score' in col else f"  Worst:   {vals.max():.2f}",
                        f"  Mean:    {vals.mean():.3f}" if 'score' in col else f"  Mean:    {vals.mean():.2f}",
                    ])

        lines.extend([
            "",
            f"TOP {top_n} MOLECULES (by {rank_by}):",
            "-" * w,
            f"{'Rank':<5} {'Name':<28} {'Vina':>8} {'CNN_Sc':>7} {'CNN_Aff':>8} {'Poses':>5}",
            "-" * w,
        ])

        rank_col = {
            "cnn_affinity": "CNN_Affinity_Rank",
            "cnn_score": "CNN_Score_Rank",
            "vina_affinity": "Vina_Rank",
            "consensus": "Consensus_Rank",
        }.get(rank_by, "Consensus_Rank")

        if rank_col in successful.columns:
            top = successful.nsmallest(top_n, rank_col)
        else:
            top = successful.head(top_n)

        for rank, (_, row) in enumerate(top.iterrows(), 1):
            vina = f"{row['vina_affinity']:.2f}" if pd.notna(row.get('vina_affinity')) else "—"
            cnn = f"{row['cnn_score']:.3f}" if pd.notna(row.get('cnn_score')) else "—"
            cnn_a = f"{row['cnn_affinity']:.2f}" if pd.notna(row.get('cnn_affinity')) else "—"
            poses = str(int(row['n_poses'])) if pd.notna(row.get('n_poses')) else "—"
            lines.append(f"{rank:<5} {row['name']:<28} {vina:>8} {cnn:>7} {cnn_a:>8} {poses:>5}")

    lines.extend(["", "=" * w])
    summary_text = "\n".join(lines)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(summary_text)
    return summary_text


# =============================================================================
# MAIN PIPELINE
# =============================================================================

def run_gnina_score_collection(
        gnina_results_csv: str,
        output_dir: str,
        molecules_csv: Optional[str] = None,
        name_column: str = "Name",
        top_n: int = 10,
        rank_by: str = "cnn_affinity",
        export_excel: bool = True,
        keep_all_poses: bool = False,
) -> Dict[str, Any]:
    """Run the complete GNINA score collection pipeline.

    When keep_all_poses=True, scans the poses directory for output SDF files,
    extracts scores for every pose, and exports gnina_all_poses.csv
    (one row per pose per molecule) for clustering module 03a.
    """
    logger.info("=" * 60)
    logger.info("GNINA SCORE COLLECTION (02c) v1.1")
    logger.info("=" * 60)

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    gnina_df = pd.read_csv(gnina_results_csv)
    logger.info(f"Loaded {len(gnina_df)} rows from {gnina_results_csv}")

    meta_df = None
    if molecules_csv and Path(molecules_csv).exists():
        meta_df = pd.read_csv(molecules_csv)
        logger.info(f"Loaded metadata: {len(meta_df)} molecules")

    logger.info("Enriching and ranking...")
    enriched = enrich_and_rank(gnina_df, meta_df, name_column, top_n, rank_by)

    # CSV
    scores_csv = output_path / "gnina_scores.csv"
    enriched.to_csv(scores_csv, index=False)
    logger.info(f"  Saved CSV: {scores_csv}")

    # --- Extract all poses from SDF files (for clustering) ---
    all_poses_csv = None
    if keep_all_poses:
        # Find poses directory: sibling of gnina_results.csv
        results_parent = Path(gnina_results_csv).parent
        poses_dir = results_parent / "poses"

        if poses_dir.exists():
            all_poses = []
            sdf_files = sorted(poses_dir.glob("*_gnina.sdf"))
            logger.info(f"  Scanning {len(sdf_files)} SDF files for all poses...")

            for sdf_file in sdf_files:
                mol_name = sdf_file.stem.replace("_gnina", "")
                poses = parse_gnina_sdf_all_poses(str(sdf_file), mol_name)
                all_poses.extend(poses)

            if all_poses:
                df_all = pd.DataFrame(all_poses)
                df_all = df_all.sort_values(
                    ["Name", "vina_affinity"], ascending=[True, True],
                    na_position="last",
                ).reset_index(drop=True)

                all_poses_csv = output_path / "gnina_all_poses.csv"
                df_all.to_csv(str(all_poses_csv), index=False)

                n_molecules = df_all["Name"].nunique()
                n_total = len(df_all)
                logger.info(
                    f"  Saved: {all_poses_csv} "
                    f"({n_total} poses from {n_molecules} molecules)")
            else:
                logger.warning("  keep_all_poses=True but no poses found in SDF files")
        else:
            logger.warning(f"  Poses dir not found: {poses_dir}")

    # Excel
    scores_xlsx = None
    if export_excel:
        scores_xlsx = output_path / "gnina_scores.xlsx"
        export_to_excel(enriched, str(scores_xlsx))

    # Summary
    summary_file = output_path / "gnina_score_summary.txt"
    generate_summary(enriched, str(summary_file), top_n, rank_by)
    logger.info(f"  Saved summary: {summary_file}")

    # JSON
    scores_json = output_path / "gnina_scores.json"
    with open(scores_json, 'w') as f:
        json.dump(enriched.to_dict(orient='records'), f, indent=2, default=str)

    successful = enriched[enriched['success'] == True]
    n_ok = len(successful)

    logger.info("")
    logger.info("=" * 60)
    logger.info("SCORE COLLECTION COMPLETE")
    logger.info("=" * 60)
    logger.info(f"Molecules scored:  {n_ok}/{len(enriched)}")
    if n_ok > 0 and 'cnn_score' in successful.columns:
        logger.info(f"CNN score range:   {successful['cnn_score'].min():.3f} to {successful['cnn_score'].max():.3f}")
    if n_ok > 0 and 'vina_affinity' in successful.columns:
        logger.info(f"Vina range:        {successful['vina_affinity'].min():.2f} to {successful['vina_affinity'].max():.2f}")

    return {
        "success": True,
        "n_total": len(enriched),
        "n_scored": n_ok,
        "n_failed": len(enriched) - n_ok,
        "scores_csv": str(scores_csv),
        "scores_xlsx": str(scores_xlsx) if scores_xlsx else None,
        "scores_json": str(scores_json),
        "all_poses_csv": str(all_poses_csv) if all_poses_csv else None,
        "summary_txt": str(summary_file),
        "dataframe": enriched,
    }


if __name__ == '__main__':
    print("GNINA Score Collector - Core Module (02c) v1.0")