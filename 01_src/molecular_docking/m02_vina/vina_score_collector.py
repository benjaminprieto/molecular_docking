#!/usr/bin/env python3
"""
Vina Score Collector - Core Module (02c)
==========================================
Collects, enriches, and exports Vina docking scores.

Reads vina_results.csv from 02b, merges with molecule metadata from 00a,
ranks by affinity, and produces final CSV + Excel outputs compatible
with dock2profile downstream.

Does NOT perform cross-engine comparison (Vina vs DOCK6) —
that responsibility belongs to dock2profile.

Location: 01_src/molecular_docking/m02_vina/vina_score_collector.py

Project: molecular_docking
Module: 02c (core)
Version: 1.0
"""

import logging
import json
from pathlib import Path
from typing import Dict, List, Optional, Any, Union

import pandas as pd
import numpy as np

logger = logging.getLogger(__name__)

# Optional: openpyxl for Excel
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.debug("openpyxl not available — Excel export disabled")


# =============================================================================
# SCORE PARSING
# =============================================================================

def load_vina_results(results_csv: str) -> pd.DataFrame:
    """
    Load Vina results from 02b output CSV.

    Expected columns: name, success, affinity, rmsd_lb, rmsd_ub,
                      n_poses, output_file, error, runtime, engine
    """
    df = pd.read_csv(results_csv)
    logger.info(f"Loaded {len(df)} rows from {results_csv}")
    return df


def load_molecule_metadata(molecules_csv: str) -> pd.DataFrame:
    """
    Load molecule metadata from 00a output CSV.

    Expected columns: Name, SMILES*, MW, LogP, HBD, HBA, TPSA, QED, etc.
    """
    df = pd.read_csv(molecules_csv)
    logger.info(f"Loaded metadata: {len(df)} molecules from {molecules_csv}")
    return df


# =============================================================================
# ENRICHMENT & RANKING
# =============================================================================

def enrich_and_rank(
        vina_df: pd.DataFrame,
        meta_df: Optional[pd.DataFrame] = None,
        name_column: str = "Name",
        top_n: int = 10,
) -> pd.DataFrame:
    """
    Merge Vina scores with molecule metadata and compute rankings.

    Steps:
      1. Merge on name
      2. Rank by Vina affinity (most negative = best = rank 1)
      3. Add rank column and percentile
      4. Flag reference/control molecules

    Args:
        vina_df:     DataFrame from 02b vina_results.csv
        meta_df:     DataFrame from 00a unique_molecules.csv (optional)
        name_column: Name column in metadata CSV
        top_n:       Number of top molecules to flag

    Returns:
        Enriched DataFrame with ranking columns
    """
    df = vina_df.copy()

    # Merge with metadata
    if meta_df is not None and name_column in meta_df.columns:
        # Avoid column collisions
        merge_cols = [c for c in meta_df.columns if c not in df.columns or c == name_column]
        df = df.merge(
            meta_df[merge_cols],
            left_on='name', right_on=name_column,
            how='left',
        )
        # Drop duplicate name column
        if name_column != 'name' and name_column in df.columns:
            df.drop(columns=[name_column], inplace=True, errors='ignore')

    # Filter to successful dockings for ranking
    successful = df['success'] == True

    # Rank by affinity (lower = better, so rank ascending)
    if 'affinity' in df.columns:
        df['Vina_Rank'] = np.nan
        df.loc[successful, 'Vina_Rank'] = (
            df.loc[successful, 'affinity']
            .rank(method='min', ascending=True)  # most negative first
            .astype(int)
        )

        # Percentile (0 = best, 100 = worst)
        n_valid = successful.sum()
        if n_valid > 0:
            df.loc[successful, 'Vina_Percentile'] = (
                (df.loc[successful, 'Vina_Rank'] - 1) / max(n_valid - 1, 1) * 100
            ).round(1)

    # Flag top N
    df['Is_Top_N'] = False
    if 'Vina_Rank' in df.columns:
        df.loc[df['Vina_Rank'] <= top_n, 'Is_Top_N'] = True

    # Flag reference/control molecules
    df['Is_Reference'] = False
    if 'is_reference' in df.columns:
        df['Is_Reference'] = df['is_reference'].fillna(False).astype(bool)
    elif 'Strategy' in df.columns:
        df.loc[df['Strategy'] == 'reference_control', 'Is_Reference'] = True

    # Sort by rank
    if 'Vina_Rank' in df.columns:
        df = df.sort_values('Vina_Rank', na_position='last').reset_index(drop=True)

    return df


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def export_to_excel(
        df: pd.DataFrame,
        output_path: str,
        sheet_name: str = "Vina_Scores",
) -> bool:
    """
    Export scores to formatted Excel file.

    Includes:
      - Column formatting (numbers, colors)
      - Conditional formatting for affinity
      - Auto-width columns

    Returns:
        True if successful, False otherwise.
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available — skipping Excel export")
        return False

    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Write header
        headers = list(df.columns)
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, (col_name, value) in enumerate(row.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Handle numpy/pandas types
                if pd.isna(value):
                    cell.value = None
                elif isinstance(value, (np.integer,)):
                    cell.value = int(value)
                elif isinstance(value, (np.floating,)):
                    cell.value = float(value)
                elif isinstance(value, (np.bool_,)):
                    cell.value = bool(value)
                else:
                    cell.value = value

                # Number formatting
                if col_name == 'affinity' and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
                elif col_name in ('MW', 'rmsd_lb', 'rmsd_ub') and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
                elif col_name in ('LogP', 'TPSA', 'QED') and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
                elif col_name == 'runtime' and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.1'

            # Highlight top molecules
            if row.get('Is_Top_N', False):
                green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = green

            # Highlight reference
            if row.get('Is_Reference', False):
                yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow

        # Auto-width columns
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row_idx in range(2, min(len(df) + 2, 50)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 40)

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(output_path)
        logger.info(f"  Saved Excel: {output_path}")
        return True

    except Exception as e:
        logger.error(f"Excel export failed: {e}")
        return False


# =============================================================================
# SUMMARY GENERATION
# =============================================================================

def generate_summary(
        df: pd.DataFrame,
        output_path: str,
        top_n: int = 10,
) -> str:
    """Generate a human-readable summary of Vina scores."""
    w = 70
    successful = df[df['success'] == True]
    n_ok = len(successful)
    n_fail = len(df) - n_ok

    lines = [
        "=" * w,
        "02c VINA SCORE COLLECTION — SUMMARY",
        "=" * w,
        "",
        f"Total molecules:   {len(df)}",
        f"Successful:        {n_ok}",
        f"Failed:            {n_fail}",
    ]

    if n_ok > 0 and 'affinity' in successful.columns:
        affs = successful['affinity'].dropna()
        if not affs.empty:
            lines.extend([
                "",
                "AFFINITY STATISTICS (kcal/mol):",
                f"  Best:    {affs.min():.2f}",
                f"  Worst:   {affs.max():.2f}",
                f"  Mean:    {affs.mean():.2f}",
                f"  Median:  {affs.median():.2f}",
                f"  Std:     {affs.std():.2f}",
            ])

        # Top N
        lines.extend([
            "",
            f"TOP {top_n} MOLECULES BY VINA AFFINITY:",
            "-" * w,
            f"{'Rank':<6} {'Name':<30} {'Affinity':>10} {'Poses':>6}",
            "-" * w,
        ])

        top = successful.nsmallest(top_n, 'affinity')
        for rank, (_, row) in enumerate(top.iterrows(), 1):
            aff = f"{row['affinity']:.2f}" if pd.notna(row.get('affinity')) else "—"
            poses = str(int(row['n_poses'])) if pd.notna(row.get('n_poses')) else "—"
            lines.append(f"{rank:<6} {row['name']:<30} {aff:>10} {poses:>6}")

    # Reference molecule
    refs = df[df.get('Is_Reference', False) == True] if 'Is_Reference' in df.columns else pd.DataFrame()
    if not refs.empty:
        lines.extend(["", "REFERENCE CONTROL:"])
        for _, row in refs.iterrows():
            aff = f"{row['affinity']:.2f}" if pd.notna(row.get('affinity')) else "—"
            rank = int(row['Vina_Rank']) if pd.notna(row.get('Vina_Rank')) else "—"
            lines.append(f"  {row['name']}: affinity={aff}, rank={rank}")

    lines.extend(["", "=" * w])

    summary_text = "\n".join(lines)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(summary_text)

    return summary_text


# =============================================================================
# MAIN PIPELINE
# =============================================================================

def run_vina_score_collection(
        vina_results_csv: str,
        output_dir: str,
        molecules_csv: Optional[str] = None,
        name_column: str = "Name",
        top_n: int = 10,
        export_excel: bool = True,
) -> Dict[str, Any]:
    """
    Run the complete Vina score collection pipeline.

    Steps:
      1. Load Vina results from 02b
      2. Load molecule metadata from 00a (optional)
      3. Merge, rank, and enrich
      4. Export to CSV + Excel

    Args:
        vina_results_csv: Path to vina_results.csv from 02b.
        output_dir:       Output directory.
        molecules_csv:    Path to unique_molecules.csv from 00a (optional).
        name_column:      Name column in molecules CSV.
        top_n:            Number of top molecules to highlight.
        export_excel:     Generate Excel file.

    Returns:
        Dict with summary statistics and output file paths.
    """
    logger.info("=" * 60)
    logger.info("VINA SCORE COLLECTION (02c) v1.0")
    logger.info("=" * 60)

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    # Load results
    vina_df = load_vina_results(vina_results_csv)

    # Load metadata
    meta_df = None
    if molecules_csv and Path(molecules_csv).exists():
        meta_df = load_molecule_metadata(molecules_csv)

    # Enrich and rank
    logger.info("Enriching and ranking...")
    enriched = enrich_and_rank(
        vina_df=vina_df,
        meta_df=meta_df,
        name_column=name_column,
        top_n=top_n,
    )

    # Export CSV
    scores_csv = output_path / "vina_scores.csv"
    enriched.to_csv(scores_csv, index=False)
    logger.info(f"  Saved CSV: {scores_csv}")

    # Export Excel
    scores_xlsx = None
    if export_excel:
        scores_xlsx = output_path / "vina_scores.xlsx"
        export_to_excel(enriched, str(scores_xlsx))

    # Generate summary
    summary_file = output_path / "vina_score_summary.txt"
    summary_text = generate_summary(enriched, str(summary_file), top_n=top_n)
    logger.info(f"  Saved summary: {summary_file}")

    # Save enriched JSON
    scores_json = output_path / "vina_scores.json"
    json_data = enriched.to_dict(orient='records')
    with open(scores_json, 'w') as f:
        json.dump(json_data, f, indent=2, default=str)

    # Statistics
    successful = enriched[enriched['success'] == True]
    n_ok = len(successful)

    logger.info("")
    logger.info("=" * 60)
    logger.info("SCORE COLLECTION COMPLETE")
    logger.info("=" * 60)
    logger.info(f"Molecules scored:  {n_ok}/{len(enriched)}")
    if n_ok > 0 and 'affinity' in successful.columns:
        affs = successful['affinity'].dropna()
        if not affs.empty:
            logger.info(f"Affinity range:    {affs.min():.2f} to {affs.max():.2f} kcal/mol")
            logger.info(f"Mean affinity:     {affs.mean():.2f} kcal/mol")

    return {
        "success": True,
        "n_total": len(enriched),
        "n_scored": n_ok,
        "n_failed": len(enriched) - n_ok,
        "scores_csv": str(scores_csv),
        "scores_xlsx": str(scores_xlsx) if scores_xlsx else None,
        "scores_json": str(scores_json),
        "summary_txt": str(summary_file),
        "dataframe": enriched,
    }


if __name__ == '__main__':
    print("Vina Score Collector - Core Module (02c) v1.0")
    print("Use 02c_vina_score_collector.py CLI for execution")